from django.views.generic import CreateView
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from datetime import date
from openpyxl import Workbook, load_workbook
from .models import UploadFile
from .forms import UploadForm
from test_app.models import TestModel


# Create your views here.
@method_decorator(login_required, name='dispatch')
class ImportView(CreateView):
    model = UploadFile
    form_class = UploadForm
    success_url = '/'
    template_name = 'import_app/import.html'

    def form_valid(self, form, **kwargs):
        o = form.save(commit=False)
        o.log = ''
        filename = o.file
        wb = load_workbook(filename=filename)
        sheet = wb.active
        if sheet.cell(row=1, column=1).value == 'Название' and sheet.cell(row=1, column=2).value == 'Описание':
            try:
                n = 2
                while sheet.cell(row=n, column=1).value:
                    post = TestModel(
                        name=sheet.cell(row=n, column=1).value,
                        text=sheet.cell(row=n, column=2).value
                    )
                    post.save()
                    o.log += f"Создан пост {sheet.cell(row=n, column=1).value}\n"
                    n += 1
                o.is_ok = True
            except Exception as e:
                o.log += str(e)
                o.is_ok = False
        else:
            o.is_ok = False
            o.log += 'Формат файла неверный\n'
        return super(ImportView, self).form_valid(form)
